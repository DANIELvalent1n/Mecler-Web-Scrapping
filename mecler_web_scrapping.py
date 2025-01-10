import streamlit as st
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import os
import re
import time
import shutil
import subprocess
from typing import List, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter 
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

def get_logpath() -> str:
    """Ensure the directory exists and return the log file path."""
    log_dir = os.path.join(os.getcwd(), 'logs')
    os.makedirs(log_dir, exist_ok=True)
    return os.path.join(log_dir, 'selenium.log')

def delete_selenium_log(logpath: str):
    """Delete the Selenium log file if it exists."""
    if os.path.exists(logpath):
        os.remove(logpath)

def get_chromedriver_path() -> str:
    """Return the path to the chromedriver executable."""
    return shutil.which('chromedriver')

def get_webdriver_options() -> Options:
    """Return configured Selenium WebDriver options."""
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-features=NetworkService")
    options.add_argument("--window-size=1920x1080")
    options.add_argument("--disable-features=VizDisplayCompositor")
    options.add_argument('--ignore-certificate-errors')
    options.set_capability('goog:loggingPrefs', {'performance': 'ALL'})
    return options

def get_webdriver_service(logpath) -> Service:
    """Create and return a Selenium WebDriver service."""
    service = Service(
        executable_path=get_chromedriver_path(),
        log_output=logpath,
    )
    return service

def validate_and_format_url(url: str) -> str:
    """Ensure the URL starts with http:// or https://, otherwise prepend https://."""
    if not url.startswith(("http://", "https://")):
        return "https://" + url
    return url

def search_and_visit_links(url: str, search_text, logpath: str) -> Tuple[str, dict, str]:
    # Convertim textul de căutare în cuvinte individuale
    keywords = search_text.lower().split()

    options = get_webdriver_options()
    service = get_webdriver_service(logpath=logpath)

    # Exemplu de utilizare
    driver = webdriver.Chrome(options=options)
    
    results = []  # Pentru stocarea rezultatelor
    links_data = []  # Vom salva aici linkurile pentru a le exporta în Excel
    with webdriver.Chrome(options=options, service=service) as driver:
        try:
            # Accesăm URL-ul
            driver.get(url)
            time.sleep(5)  # Așteptăm ca pagina să se încarce complet

            # Căutăm bara de căutare
            search_box = None
            possible_search_selectors = [
                {"by": By.NAME, "value": "q"}, 
                {"by": By.ID, "value": "search"},
                {"by": By.CLASS_NAME, "value": "search"},
                {"by": By.TAG_NAME, "value": "input"}
            ]
            
            for selector in possible_search_selectors:
                try:
                    search_box = driver.find_element(selector["by"], selector["value"])
                    if search_box:
                        break
                except:
                    pass

            if not search_box:
                results.append("Bara de căutare nu a fost găsită pe acest site.")
                return results, None

            # Introducem textul în bara de căutare și apăsăm Enter
            search_box.send_keys(search_text)
            search_box.send_keys(Keys.RETURN)
            time.sleep(3)  # Așteptăm să se încarce rezultatele

            # Găsim toate linkurile din pagină
            all_links = driver.find_elements(By.TAG_NAME, "a")
            filtered_links = []

            # Filtrăm linkurile relevante
            for link_element in all_links:
                href = link_element.get_attribute("href")
                if href:
                    # Filtrare pentru olx sau autovit
                    if ("olx" in url or "autovit" in url or "publi24" in url) and ("/d/oferta" in href or "/anunt" in href or "/anunturi" in href) and any(keyword in href.lower() for keyword in keywords) and '.html' in href.lower():
                        filtered_links.append(href)
            
            # Eliminăm linkurile duplicate
            filtered_links = list(set(filtered_links))
            
            results.append(f"Am găsit {len(filtered_links)} linkuri relevante:")
            results.extend(filtered_links)
            
            # Navigăm pe primele 5 linkuri
            for link in filtered_links[:links_number]:
                driver.get(link)
                time.sleep(2)  # Așteptăm să se încarce pagina

                # Căutăm primul div care conține "description" în oricare atribut
                description_div = None
                try:
                    if "publi24" in url:
                        description_div = driver.find_element(By.XPATH, "//*[contains(@itemprop, 'description')]")
                    else:
                        description_div = driver.find_element(By.XPATH, "//*[contains(@*, 'description')]")
                except:
                    print(f"Nu am găsit un div cu 'description' pentru link-ul {link}")
                    continue

                # Extragem titlul anunțului (de obicei, titlul este în tag-ul <title>)
                title = driver.title  # Titlul paginii este adesea titlul anunțului

                # Extragem textul din div-ul care conține "description"
                description_text = description_div.get_attribute("innerText").strip()

                # Salvăm informațiile selectate
                entry = {}
                if save_link:
                    entry["Link"] = link
                if save_title:
                    entry["Titlu"] = title
                if save_description:
                    entry["Descriere"] = description_text

                links_data.append(entry)
        finally:
            # Închidem browser-ul
            driver.quit()

    return results, links_data

def save_to_excel_with_wrap(links_data):
    # Creăm un DataFrame din datele linkurilor
    df = pd.DataFrame(links_data)

    # Salvăm DataFrame-ul într-un fișier Excel
    excel_file = "lista_informatii_extrase.xlsx"
    df.to_excel(excel_file, index=False)

    # Deschidem fișierul Excel cu openpyxl pentru a aplica formatul "wrap-text"
    wb = load_workbook(excel_file)
    ws = wb.active

    # Calculăm lățimea pentru fiecare coloană
    num_columns = 3  # Numărul total de coloane (Titlu, Descriere, etc.)
    column_width = (0.9 * 100) / num_columns  # Distribuim egal

    # Setăm lățimea pentru fiecare coloană
    for col_idx in range(1, 4):  # Coloanele 2 și 3
        col_letter = get_column_letter(col_idx)  # Obținem litera coloanei (ex: A, B, etc.)
        ws.column_dimensions[col_letter].width = column_width

    # Aplicăm wrap text pentru celule
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):  # Col 2 este pentru "Titlu" și Col 3 pentru "Descriere"
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
    
    # Salvăm fișierul cu formatarea aplicată
    wb.save(excel_file)

    return excel_file

# Interfața Streamlit
st.title("Căutare avansată pe site-uri cu Mecler Web Scraper")

# Introducerea URL-ului și a textului de căutare
url = st.text_input("Introduceți URL-ul site-ului:", placeholder="Ex: https://www.olx.ro/")
search_text = st.text_input("Introduceți textul pentru căutare:", placeholder="Ex: Nissan Qashqai 2010")
links_number = st.text_input("Introduceți numarul de anunturi dorit pentru analizare:", placeholder="Ex: 14")

# Verificăm dacă valoarea introdusă este un număr valid
if links_number.isdigit():
    links_number = int(links_number)  # Convertim în int

# Selectarea criteriilor de salvare
st.write("Selectați informațiile pe care doriți să le extrageti din anunturi:")
save_title = st.checkbox("Salvează titlul", value=True)
save_link = st.checkbox("Salvează linkul", value=True)
save_description = st.checkbox("Salvează descrierea", value=True)

# Buton pentru inițierea căutării
if st.button("Extrage date"):
    with st.spinner("Procesăm cererea, vă rugăm să așteptați..."):
        if url and search_text and links_number:

            logpath = get_logpath()
            delete_selenium_log(logpath=logpath)
            
            results, links_data = search_and_visit_links(url, search_text, logpath)
            # Verificăm dacă sunt linkuri disponibile pentru salvare
            if links_data:
                # Salvăm linkurile, titlurile și descrierile în Excel cu wrap text activat
                excel_file = save_to_excel_with_wrap(links_data)

                # Oferim fișierul pentru descărcare
                with open(excel_file, "rb") as file:
                    st.download_button(
                        label="Descarcă rezultatele în Excel",
                        data=file,
                        file_name=excel_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            else:
                st.write("Nu au fost găsite descriere relevante.")
        else:
            st.error("Vă rugăm să completați atât URL-ul, textul, cat si numarul pentru căutare.")
